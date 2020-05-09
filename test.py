
from pydriller import RepositoryMining
import subprocess

import re
import json
import os

import csv

import xlsxwriter

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import openpyxl

import shutil

ApplicationName = "laravel"

pathDirectory = "D:/Projects/AnalysteProject/repositories/" + ApplicationName

os.chdir("D:/Projects/AnalysteProject/repositories")


Projet_GIT_URL = "https://github.com/laravel/laravel"

# subprocess.check_output("git clone " + Projet_GIT_URL , shell=True)

os.chdir("D:/Projects/AnalysteProject")

f = open('D:/Projects/AnalysteProject/config_' + ApplicationName +'.json')
json_file = json.load(f)
print(json_file)
if not json_file['name']:
    # workbook = xlsxwriter.Workbook("D:/Projects/AnalysteProject/" + ApplicationName + "/Analyse_" + ApplicationName + ".xlsx")
    # worksheet = workbook.add_worksheet("Analyse")
    # wsn = worksheet.name
    # print(worksheet.name)
    row = 2
    Files = []
    start_commit = json_file['first_commit']
    # worksheet.write('A1', 'Commit id')
    # worksheet.write('B1', 'Date')
    # worksheet.write('C1', 'filename')
    # worksheet.write('D1', 'CyclomaticComplexity')
    # worksheet.write('E1', 'ExcessiveClassLength')
    # worksheet.write('F1', 'ExcessiveMethodLength')
    # worksheet.write('G1', 'ExcessiveParameterList')
    # worksheet.write('H1', 'NPathComplexity')
    # worksheet.write('I1', 'CouplingBetweenObjects')
    # worksheet.write('J1', 'EmptyCatchBlock')
    # worksheet.write('K1', 'DepthOfInheritance')
    # worksheet.write('L1', 'GotoStatement')
    # workbook.close()

    csvfile = open("D:/Projects/AnalysteProject/" + ApplicationName + "/Analyse_" + ApplicationName  + ".csv", 'w', newline='')

    with csvfile:

        writer = csv.writer(csvfile, delimiter='|')

        writer.writerow(('Commit id', 'Date', 'filename',
                         'CyclomaticComplexity', 'ExcessiveClassLength',
                         'ExcessiveMethodLength', 'ExcessiveParameterList', 'NPathComplexity', 'CouplingBetweenObjects',
                         'EmptyCatchBlock', 'DepthOfInheritance', 'GotoStatement'))
    csvfile.close()
else:
    row = json_file['row']
    start_commit = json_file['commit']
    Files = json_file['files']


f.close()
# json_file['config'] = {'name': "D:/Projects/AnalysteProject/" + ApplicationName + "/Analyse_" + ApplicationName + ".xlsx",
#                        'row': 1,
#                        'commit': '0'}

# Use the worksheet object to write
# data via the write() method.







def smell_cmd(commit, row, commit_date):
    try:

        # # wb = load_workbook(filename="D:/Projects/AnalysteProject/" + ApplicationName + "/Analyse_" + ApplicationName + ".xlsx")
        # # workbook = wb.active
        # wbkName = "D:/Projects/AnalysteProject/" + ApplicationName + "/Analyse_" + ApplicationName + ".xlsx"
        # wbk = openpyxl.load_workbook(wbkName)
        # # print(wbk.worksheets[0])
        # wks = wbk.worksheets[0]


        out = subprocess.check_output("phpmd " + pathDirectory + " json D:/Projects/AnalysteProject/myRuleset.xml ", shell=True)
    except subprocess.CalledProcessError as e:
        out = e.output

    result_dict = json.loads(out)
    # print(result_dict)

    date_time = commit_date.strftime("%m/%d/%Y, %H:%M:%S")

    copy_files = Files.copy()
    # print('Files:', len(result_dict['files']))
    for file in result_dict['files']:

        if file['file'] not in Files:
            Files.append(file['file'])
        else:
            copy_files.remove(file['file'])


        NPathComplexity = 0
        CyclomaticComplexity = 0
        ExcessiveClassLength = 0
        ExcessiveMethodLength =0
        ExcessiveParameterList = 0
        CouplingBetweenObjects = 0
        EmptyCatchBlock = 0
        DepthOfInheritance = 0
        GotoStatement = 0

        for violation in file['violations']:

            if violation['rule'] == 'CyclomaticComplexity':
                CyclomaticComplexity = CyclomaticComplexity + 1
            if violation['rule'] == 'NPathComplexity':
                NPathComplexity = NPathComplexity + 1
            if violation['rule'] == 'ExcessiveClassLength':
                ExcessiveClassLength = ExcessiveClassLength + 1
            if violation['rule'] == 'ExcessiveMethodLength':
                ExcessiveMethodLength = ExcessiveMethodLength + 1
            if violation['rule'] == 'ExcessiveParameterList':
                ExcessiveParameterList = ExcessiveParameterList + 1
            if violation['rule'] == 'CouplingBetweenObjects':
                CouplingBetweenObjects = CouplingBetweenObjects + 1
            if violation['rule'] == 'EmptyCatchBlock':
                EmptyCatchBlock = EmptyCatchBlock + 1
            if violation['rule'] == 'DepthOfInheritance':
                DepthOfInheritance = DepthOfInheritance + 1
            if violation['rule'] == 'GotoStatement':
                GotoStatement = GotoStatement + 1

        # print('CodeSmells Complexite', CyclomaticComplexity)


        print(Files)
        print(copy_files)

        csvfile1 = open("D:/Projects/AnalysteProject/" + ApplicationName + "/Analyse_" + ApplicationName + 'csv', 'a', newline='')

        # with csvfile1:
        #
        #     writer = csv.writer(csvfile1, delimiter=',')
        #
        #     writer.writerow((commit, date_time, file['file'],
        #                      CyclomaticComplexity, ExcessiveClassLength,
        #                      ExcessiveMethodLength, ExcessiveParameterList, NPathComplexity, CouplingBetweenObjects,
        #                      EmptyCatchBlock, DepthOfInheritance, GotoStatement))
        # csvfile1.close()

        # worksheet.write_row(row, 0, commit)
        # worksheet.write(row, 0, commit)
        # worksheet.write(row, 1, date_time)
        # worksheet.write(row, 2, file['file'])
        # worksheet.write(row, 3, CyclomaticComplexity)
        # worksheet.write(row, 4, ExcessiveClassLength)
        # worksheet.write(row, 5, ExcessiveMethodLength)
        # worksheet.write(row, 6, ExcessiveParameterList)
        # worksheet.write(row, 7, NPathComplexity)
        # worksheet.write(row, 8, CouplingBetweenObjects)
        # worksheet.write(row, 9, EmptyCatchBlock)
        # worksheet.write(row, 10, DepthOfInheritance)
        # worksheet.write(row, 11, GotoStatement)

        # incrementing the value of row by one
        # with each iteratons.
        row += 1

    for elemet_file in copy_files:

        # EXCEL Works START
        # print(elemet_file)
        # wks.cell(row=row, column=1).value = commit
        # wks.cell(row=row, column=2).value =date_time
        # wks.cell(row=row, column=3).value = elemet_file
        # wks.cell(row=row, column=4).value = 0
        # wks.cell(row=row, column=5).value = 0
        # wks.cell(row=row, column=6).value = 0
        # wks.cell(row=row, column=7).value = 0
        # wks.cell(row=row, column=8).value = 0
        # wks.cell(row=row, column=9).value = 0
        # wks.cell(row=row, column=10).value = 0
        # wks.cell(row=row, column=11).value = 0
        # wks.cell(row=row, column=12).value = 0
        # EXCEL END

        row += 1

        # wbk.save(wbkName)
        csvfile1 = open("D:/Projects/AnalysteProject/" + ApplicationName + "/Analyse_" + ApplicationName + 'csv', 'a', newline='')

        # with csvfile1:
        #     writer = csv.writer(csvfile1, delimiter=',')
        #     writer.writerow((commit, date_time, elemet_file, 0, 0, 0, 0, 0, 0, 0, 0, 0))
        # csvfile1.close()

    # wbk.close

    return row






count = 0


os.chdir(pathDirectory)
for commit in RepositoryMining(pathDirectory, from_commit=start_commit).traverse_commits():
    count += 1
    print("Commit :", count)

    cmd_Checkout = "git checkout " + commit.hash + " -f"
    # print(cmd_Checkout)
    # print('CLEAN')
    # os.system("git reset --hard")
    print('Checkout !!!!!')
    subprocess.check_output(cmd_Checkout, shell=True)
    # os.system(cmd_Checkout)


#     json_file = {
#         'name': "D:/Projects/AnalysteProject/" + ApplicationName + "/Analyse_" + ApplicationName + ".csv",
#         'row': row,
#         'commit': commit.hash,
#         'first_commit': "a188d62105532fcf2a2839309fb71b862d904612",
#         'files': Files
#     }
#
#     with open('D:/Projects/AnalysteProject/config_' + ApplicationName +'.json', 'w') as gg:
#         # print(json_file)
#         gg.seek(0)
#         json.dump(json_file, gg)
#
    # row = smell_cmd(commit.hash, row, commit.committer_date)
#
#     json_file = {
#         'name': "D:/Projects/AnalysteProject/" + ApplicationName + "/Analyse_" + ApplicationName + ".csv",
#         'row': row,
#         'commit': commit.hash,
#         'first_commit': "a188d62105532fcf2a2839309fb71b862d904612",
#         'files': Files
#     }
#     with open('D:/Projects/AnalysteProject/config_' + ApplicationName + '.json', 'w') as gg:
#         print(json_file)
#         gg.seek(0)
#         json.dump(json_file, gg)
#
#     f = open('D:/Projects/AnalysteProject/config_' + ApplicationName +'.json')
#     json_file = json.load(f)
#     # print(json_file)
#     f.close()
#
print("Closed")
#
# f = open('D:/Projects/AnalysteProject/config_' + ApplicationName +'.json')
# json_file = json.load(f)
# print(json_file["row"])
# f.close()
