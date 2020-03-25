
from pydriller import RepositoryMining
import subprocess

import re
import json
import os

import xlsxwriter

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import openpyxl

import shutil

ApplicationName = "moodle"

pathDirectory = "D:/Projects/AnalysteProject/repositories/" + ApplicationName

os.chdir("D:/Projects/AnalysteProject/repositories")


Projet_GIT_URL = "https://github.com/moodle/moodle"

# subprocess.check_output("git clone " + Projet_GIT_URL , shell=True)

os.chdir("D:/Projects/AnalysteProject")

f = open('D:/Projects/AnalysteProject/config.json')
json_file = json.load(f)
print(json_file)
if not json_file['name']:
    workbook = xlsxwriter.Workbook("D:/Projects/AnalysteProject/" + ApplicationName + "/Analyse_" + ApplicationName + ".xlsx")
    worksheet = workbook.add_worksheet("Analyse")
    wsn = worksheet.name
    print(worksheet.name)
    row = 2
    Files = []
    start_commit = json_file['first_commit']
    worksheet.write('A1', 'Commit id')
    worksheet.write('B1', 'Date')
    worksheet.write('C1', 'filename')
    worksheet.write('D1', 'CyclomaticComplexity')
    worksheet.write('E1', 'ExcessiveClassLength')
    worksheet.write('F1', 'ExcessiveMethodLength')
    worksheet.write('G1', 'ExcessiveParameterList')
    worksheet.write('H1', 'NPathComplexity')
    worksheet.write('I1', 'CouplingBetweenObjects')
    worksheet.write('J1', 'EmptyCatchBlock')
    worksheet.write('K1', 'DepthOfInheritance')
    worksheet.write('L1', 'GotoStatement')
    workbook.close()
else:
    row = json_file['row']
    start_commit = json_file['commit']
    Files = json_file['files']


f.close()
# json_file['config'] = {'name': "D:/Projects/AnalysteProject/" + ApplicationName + "/Analyse_" + ApplicationName + ".xlsx",
#                        'row': 1,
#                        'commit': '0'}

# Use the worksheet object to write
# data via the write() method.







def smell_cmd(commit, row, commit_date):
    try:

        # wb = load_workbook(filename="D:/Projects/AnalysteProject/" + ApplicationName + "/Analyse_" + ApplicationName + ".xlsx")
        # workbook = wb.active
        wbkName = "D:/Projects/AnalysteProject/" + ApplicationName + "/Analyse_" + ApplicationName + ".xlsx"
        wbk = openpyxl.load_workbook(wbkName)
        # print(wbk.worksheets[0])
        wks = wbk.worksheets[0]


        out = subprocess.check_output("phpmd " + pathDirectory + " json D:/Projects/AnalysteProject/myRuleset.xml ", shell=True)
    except subprocess.CalledProcessError as e:
        out = e.output

    result_dict = json.loads(out)
    # print(result_dict)

    copy_files = Files.copy()
    # print('Files:', len(result_dict['files']))
    for file in result_dict['files']:

        if file['file'] not in Files:
            Files.append(file['file'])
            for elem in Files:
                print(elem)
        else:
            copy_files.remove(file['file'])
            print(copy_files)

        NPathComplexity = 0
        CyclomaticComplexity = 0
        ExcessiveClassLength = 0
        ExcessiveMethodLength =0
        ExcessiveParameterList = 0
        CouplingBetweenObjects = 0
        EmptyCatchBlock = 0
        DepthOfInheritance = 0
        GotoStatement = 0

        for violation in file['violations']:

            if violation['rule'] == 'CyclomaticComplexity':
                CyclomaticComplexity = CyclomaticComplexity + 1
            if violation['rule'] == 'NPathComplexity':
                NPathComplexity = NPathComplexity + 1
            if violation['rule'] == 'ExcessiveClassLength':
                ExcessiveClassLength = ExcessiveClassLength + 1
            if violation['rule'] == 'ExcessiveMethodLength':
                ExcessiveMethodLength = ExcessiveMethodLength + 1
            if violation['rule'] == 'ExcessiveParameterList':
                ExcessiveParameterList = ExcessiveParameterList + 1
            if violation['rule'] == 'CouplingBetweenObjects':
                CouplingBetweenObjects = CouplingBetweenObjects + 1
            if violation['rule'] == 'EmptyCatchBlock':
                EmptyCatchBlock = EmptyCatchBlock + 1
            if violation['rule'] == 'DepthOfInheritance':
                DepthOfInheritance = DepthOfInheritance + 1
            if violation['rule'] == 'GotoStatement':
                GotoStatement = GotoStatement + 1

        # print('CodeSmells Complexite', CyclomaticComplexity)

        date_time = commit_date.strftime("%m/%d/%Y, %H:%M:%S")

        # write operation perform

        wks.cell(row=row, column=1).value = commit
        wks.cell(row=row, column=2).value =date_time
        wks.cell(row=row, column=3).value = file['file']
        wks.cell(row=row, column=4).value = CyclomaticComplexity
        wks.cell(row=row, column=5).value = ExcessiveClassLength
        wks.cell(row=row, column=6).value = ExcessiveMethodLength
        wks.cell(row=row, column=7).value = ExcessiveParameterList
        wks.cell(row=row, column=8).value = NPathComplexity
        wks.cell(row=row, column=9).value = CouplingBetweenObjects
        wks.cell(row=row, column=10).value = EmptyCatchBlock
        wks.cell(row=row, column=11).value = DepthOfInheritance
        wks.cell(row=row, column=12).value = GotoStatement

        wbk.save(wbkName)


        # worksheet.write_row(row, 0, commit)
        # worksheet.write(row, 0, commit)
        # worksheet.write(row, 1, date_time)
        # worksheet.write(row, 2, file['file'])
        # worksheet.write(row, 3, CyclomaticComplexity)
        # worksheet.write(row, 4, ExcessiveClassLength)
        # worksheet.write(row, 5, ExcessiveMethodLength)
        # worksheet.write(row, 6, ExcessiveParameterList)
        # worksheet.write(row, 7, NPathComplexity)
        # worksheet.write(row, 8, CouplingBetweenObjects)
        # worksheet.write(row, 9, EmptyCatchBlock)
        # worksheet.write(row, 10, DepthOfInheritance)
        # worksheet.write(row, 11, GotoStatement)

        # incrementing the value of row by one
        # with each iteratons.
        row += 1

    for elemet_file in copy_files:
        # worksheet.write(row, 0, commit)
        # worksheet.write(row, 1, date_time)
        # worksheet.write(row, 2, elemet_file)
        # worksheet.write(row, 3, 0)
        # worksheet.write(row, 4, 0)
        # worksheet.write(row, 5, 0)
        # worksheet.write(row, 6, 0)
        # worksheet.write(row, 7, 0)
        # worksheet.write(row, 8, 0)
        # worksheet.write(row, 9, 0)
        # worksheet.write(row, 10, 0)
        # worksheet.write(row, 11, 0)
        print(elemet_file)
        wks.cell(row=row, column=1).value = commit
        wks.cell(row=row, column=2).value =date_time
        wks.cell(row=row, column=3).value = elemet_file
        wks.cell(row=row, column=4).value = 0
        wks.cell(row=row, column=5).value = 0
        wks.cell(row=row, column=6).value = 0
        wks.cell(row=row, column=7).value = 0
        wks.cell(row=row, column=8).value = 0
        wks.cell(row=row, column=9).value = 0
        wks.cell(row=row, column=10).value = 0
        wks.cell(row=row, column=11).value = 0
        wks.cell(row=row, column=12).value = 0

        row += 1

        wbk.save(wbkName)
    wbk.close

    return row






count = 0


os.chdir(pathDirectory)
for commit in RepositoryMining(pathDirectory, from_commit=start_commit).traverse_commits():
    count += 1
    print("Commit :", count)

    cmd_Checkout = "git checkout " + commit.hash
    # print(cmd_Checkout)
    subprocess.check_output(cmd_Checkout, shell=True)
    # os.system(cmd_Checkout)
    print("start")

    json_file = {
        'name': "D:/Projects/AnalysteProject/" + ApplicationName + "/Analyse_" + ApplicationName + ".xlsx",
        'row': row,
        'commit': commit.hash,
        'first_commit': "f9903ed0a41ce4df0cb3628a06d6c0a9455ac75c",
        'files': Files
    }

    with open('D:/Projects/AnalysteProject/config.json', 'w') as gg:
        # print(json_file)
        gg.seek(0)
        json.dump(json_file, gg)

    row = smell_cmd(commit.hash, row, commit.committer_date)

    json_file = {
        'name': "D:/Projects/AnalysteProject/" + ApplicationName + "/Analyse_" + ApplicationName + ".xlsx",
        'row': row,
        'commit': commit.hash,
        'first_commit': "f9903ed0a41ce4df0cb3628a06d6c0a9455ac75c",
        'files': Files
    }
    with open('D:/Projects/AnalysteProject/config.json', 'w') as gg:
        print(json_file)
        gg.seek(0)
        json.dump(json_file, gg)

    f = open('D:/Projects/AnalysteProject/config.json')
    json_file = json.load(f)
    # print(json_file)
    f.close()

print("Closed")

f = open('D:/Projects/AnalysteProject/config.json')
json_file = json.load(f)
print(json_file["row"])
f.close()
